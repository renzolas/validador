import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.comments import Comment
import io
import difflib

# ======================
# CONFIGURACIÓN STREAMLIT
# ======================
st.set_page_config(page_title="Validador de Archivos", layout="centered")
st.title("📊 Validador de coincidencias entre archivos Excel")

# ======================
# FUNCIONES AUXILIARES
# ======================
def validar_extension(nombre_archivo: str) -> bool:
    return nombre_archivo.lower().endswith(('.xlsx', '.xlsm'))

def comparar_y_resaltar(
    archivo_a,
    archivo_b,
    color_diferencia="FF9999",
    color_insert="FFFF99",
    agregar_comentarios=True,
    normalize_for_alignment=True,
    header=True
) -> io.BytesIO:
    """
    Compara archivo A (referencia) vs archivo B (a modificar) alineando filas.
    - Se usa difflib.SequenceMatcher sobre tuplas de fila para detectar insert/delete/replace.
    - Las filas añadidas en B se colorean con color_insert.
    - Las celdas distintas se colorean con color_diferencia y reciben comentario si se pide.
    - Se añade hoja "Resumen_filas_faltantes" con filas que están en A pero no en B.
    """

    # Leer con pandas (las columnas del Excel se usan como encabezados)
    df_a = pd.read_excel(archivo_a, dtype=str, engine="openpyxl").fillna("")
    df_b = pd.read_excel(archivo_b, dtype=str, engine="openpyxl").fillna("")

    # Copias originales para mostrar en comentarios (sin normalizar)
    orig_a = df_a.copy()
    orig_b = df_b.copy()

    # Definir número máximo de columnas a comparar (completa con "")
    max_cols = max(len(df_a.columns), len(df_b.columns))

    # Normalización para alineado: opcional (strip + lower)
    def norm_cell(val):
        s = "" if val is None else str(val)
        return s.strip().lower() if normalize_for_alignment else s

    # Construir listas de tuplas (una tupla por fila) para SequenceMatcher
    rows_a = []
    for r in range(len(df_a)):
        row = tuple(norm_cell(df_a.iat[r, c]) if c < len(df_a.columns) else "" for c in range(max_cols))
        rows_a.append(row)

    rows_b = []
    for r in range(len(df_b)):
        row = tuple(norm_cell(df_b.iat[r, c]) if c < len(df_b.columns) else "" for c in range(max_cols))
        rows_b.append(row)

    # Alinear con SequenceMatcher
    sm = difflib.SequenceMatcher(a=rows_a, b=rows_b)
    opcodes = sm.get_opcodes()

    # Abrir el workbook B para modificar
    wb = load_workbook(archivo_b)
    ws = wb.active

    fill_diff = PatternFill(start_color=color_diferencia, end_color=color_diferencia, fill_type="solid")
    fill_insert = PatternFill(start_color=color_insert, end_color=color_insert, fill_type="solid")

    missing_rows = []  # filas presentes en A pero ausentes en B

    # Offset para transformar índice de DataFrame -> fila Excel
    excel_offset = 2 if header else 1

    for tag, i1, i2, j1, j2 in opcodes:
        if tag == "equal":
            # Filas emparejadas (posiblemente después de normalizar).
            # Comparamos los valores originales por celda para destacar diferencias exactas.
            for a_idx, b_idx in zip(range(i1, i2), range(j1, j2)):
                for col in range(max_cols):
                    val_a = orig_a.iat[a_idx, col] if col < len(orig_a.columns) else ""
                    val_b = orig_b.iat[b_idx, col] if col < len(orig_b.columns) else ""
                    if str(val_a) != str(val_b):
                        excel_row = b_idx + excel_offset
                        cell = ws.cell(row=excel_row, column=col + 1)
                        cell.fill = fill_diff
                        if agregar_comentarios:
                            cell.comment = Comment(f'Se esperaba "{val_a}" y se encontró "{val_b}"', "Validador")
        elif tag == "replace":
            # Rango reemplazado: emparejar hasta min y tratar extras como insert/delete
            len_a = i2 - i1
            len_b = j2 - j1
            min_len = min(len_a, len_b)

            # Comparar filas emparejadas
            for k in range(min_len):
                a_idx = i1 + k
                b_idx = j1 + k
                for col in range(max_cols):
                    val_a = orig_a.iat[a_idx, col] if col < len(orig_a.columns) else ""
                    val_b = orig_b.iat[b_idx, col] if col < len(orig_b.columns) else ""
                    if str(val_a) != str(val_b):
                        excel_row = b_idx + excel_offset
                        cell = ws.cell(row=excel_row, column=col + 1)
                        cell.fill = fill_diff
                        if agregar_comentarios:
                            cell.comment = Comment(f'Se esperaba "{val_a}" y se encontró "{val_b}"', "Validador")

            # Extras en B -> insertadas
            if len_b > len_a:
                for b_idx in range(j1 + min_len, j2):
                    for col in range(max_cols):
                        excel_row = b_idx + excel_offset
                        cell = ws.cell(row=excel_row, column=col + 1)
                        cell.fill = fill_insert
                        if agregar_comentarios:
                            val_b = orig_b.iat[b_idx, col] if col < len(orig_b.columns) else ""
                            cell.comment = Comment(f'Fila añadida en B. Valor: "{val_b}"', "Validador")
            # Extras en A -> faltantes en B
            if len_a > len_b:
                for a_idx in range(i1 + min_len, i2):
                    missing_rows.append((a_idx + excel_offset, [orig_a.iat[a_idx, c] if c < len(orig_a.columns) else "" for c in range(max_cols)]))
        elif tag == "delete":
            # Filas en A que no existen en B
            for a_idx in range(i1, i2):
                missing_rows.append((a_idx + excel_offset, [orig_a.iat[a_idx, c] if c < len(orig_a.columns) else "" for c in range(max_cols)]))
        elif tag == "insert":
            # Filas insertadas en B (no están en A)
            for b_idx in range(j1, j2):
                for col in range(max_cols):
                    excel_row = b_idx + excel_offset
                    cell = ws.cell(row=excel_row, column=col + 1)
                    cell.fill = fill_insert
                    if agregar_comentarios:
                        val_b = orig_b.iat[b_idx, col] if col < len(orig_b.columns) else ""
                        cell.comment = Comment(f'Fila añadida en B. Valor: "{val_b}"', "Validador")

    # Agregar hoja resumen si faltan filas de A
    if missing_rows:
        sheet_name = "Resumen_filas_faltantes"
        if sheet_name in wb.sheetnames:
            # reemplazamos si existe
            wb.remove(wb[sheet_name])
        sum_ws = wb.create_sheet(sheet_name)
        # Encabezado
        sum_ws.append(["Fila en A (Excel)", "Contenido (columnas)"])
        for rownum, vals in missing_rows:
            sum_ws.append([rownum, " | ".join(map(str, vals))])

    # Guardar en BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# ======================
# INTERFAZ STREAMLIT
# ======================
archivo_a = st.file_uploader("📂 Sube el Archivo A (referencia)", type=["xlsx", "xlsm"])
archivo_b = st.file_uploader("📂 Sube el Archivo B (comparar y modificar)", type=["xlsx", "xlsm"])

# Opciones útiles
normalize_checkbox = st.checkbox("🔧 Ignorar mayúsculas y espacios al alinear filas (más tolerante)", value=True)
header_checkbox = st.checkbox("📑 La primera fila es encabezado (offset Excel = 2)", value=True)
comments_checkbox = st.checkbox("💬 Agregar comentarios en celdas con diferencias", value=True)

if archivo_a and archivo_b:
    if validar_extension(archivo_a.name) and validar_extension(archivo_b.name):
        if st.button("🔍 Validar Archivos"):
            try:
                resultado = comparar_y_resaltar(
                    archivo_a,
                    archivo_b,
                    color_diferencia="FF9999",
                    color_insert="FFFF99",
                    agregar_comentarios=comments_checkbox,
                    normalize_for_alignment=normalize_checkbox,
                    header=header_checkbox
                )
                st.success("✅ Comparación completada")
                st.download_button(
                    label="📥 Descargar archivo B validado",
                    data=resultado,
                    file_name="archivo_B_validado.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            except Exception as e:
                st.error(f"❌ Ocurrió un error durante la comparación: {e}")
    else:
        st.error("❌ Solo se permiten archivos con extensión .xlsx o .xlsm")

